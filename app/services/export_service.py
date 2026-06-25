"""Excel export of contacts (feature #5 daily report, #3 social profile link)."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.contact import Contact

HEADERS = [
    "الاسم",
    "رقم الهاتف",
    "الإيميل",
    "المصدر",
    "رابط الحساب",
    "الحالة",
    "التقييم",
    "عميل ساخن؟",
    "التاجات",
    "آخر نشاط",
    "تاريخ الإضافة",
]


def _profile_link(contact: Contact) -> str:
    refs = contact.external_refs if isinstance(contact.external_refs, dict) else {}
    return (
        refs.get("profile_url")
        or refs.get("profile_link")
        or refs.get("permalink")
        or ""
    )


async def build_contacts_xlsx(session: AsyncSession) -> bytes:
    """Build an .xlsx of all active contacts for the current tenant (RLS-scoped)."""
    rows = (
        await session.execute(
            select(Contact).where(Contact.deleted_at.is_(None)).order_by(Contact.created_at.desc())
        )
    ).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for c in rows:
        ws.append(
            [
                c.full_name,
                c.phone or "",
                c.email or "",
                c.source.value if c.source else "",
                _profile_link(c),
                c.status.value if c.status else "",
                int(c.lead_score or 0),
                "نعم" if (c.lead_score or 0) >= 70 else "لا",
                ", ".join(c.tags or []),
                c.last_activity_at.strftime("%Y-%m-%d %H:%M") if c.last_activity_at else "",
                c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            ]
        )

    # Reasonable column widths.
    for i, _ in enumerate(HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
